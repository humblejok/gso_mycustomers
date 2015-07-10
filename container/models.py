import datetime
import logging
import os
import traceback

from django.contrib.auth.models import User
from django.db import models
from django.db.models import Q
from django.db.models.fields import FieldDoesNotExist
from django.template import loader
from django.template.context import Context
from openpyxl.reader.excel import load_workbook
from seq_common.utils import classes

import container
from container.utilities.utils import get_static_fields, \
    complete_fields_information
from gso_mycustomers.settings import TEMPLATE_DIRS
from container.setup.application.settings import RESOURCES_MAIN_PATH,\
    TEMPLATES_STATICS_PATH


LOGGER = logging.getLogger(__name__)

def setup():
    setup_attributes()
    setup_labels()
    populate_model_from_xlsx('container.models.MenuEntries', os.path.join(RESOURCES_MAIN_PATH,'Repository Setup.xlsx'))
    generate_attributes()
    generate_wizards()
    setup_menus()
    setup_global_menus()
def setup_menus():
    MenuEntries.objects.all().delete()
    populate_model_from_xlsx('container.models.MenuEntries', os.path.join(RESOURCES_MAIN_PATH,'Repository Setup.xlsx'))
    GlobalMenuEntries.objects.all().delete()
    populate_model_from_xlsx('container.models.GlobalMenuEntries', os.path.join(RESOURCES_MAIN_PATH,'Repository Setup.xlsx'))
    generate_global_templates()
    
def setup_global_menus():
    GlobalMenuEntries.objects.all().delete()
    populate_model_from_xlsx('container.models.GlobalMenuEntries', os.path.join(RESOURCES_MAIN_PATH,'Repository Setup.xlsx'))
    generate_global_templates()

def setup_labels():
    FieldLabel.objects.all().delete()
    populate_labels_from_xlsx('container.models.FieldLabel', os.path.join(RESOURCES_MAIN_PATH,'Repository Setup.xlsx'))

def setup_attributes():
    populate_attributes_from_xlsx('container.models.Attributes', os.path.join(RESOURCES_MAIN_PATH,'Repository Setup.xlsx'))
    populate_attributes_from_xlsx('container.models.Dictionary', os.path.join(RESOURCES_MAIN_PATH,'Repository Setup.xlsx'))

def generate_wizards():
    wizards = Attributes.objects.filter(active=True, type='element_wizard')
    languages = Attributes.objects.filter(active=True, type='available_language')
    template = loader.get_template('rendition/container_type/creations/wizard.html')
    for language in languages:
        for wizard in wizards:
            working_class = classes.my_class_import(wizard.name)
            all_fields = get_static_fields(working_class)
            all_fields = complete_fields_information(working_class, all_fields, language.short_name)
            context = Context({'fields': working_class.get_wizard_fields(), 'complete_fields': all_fields, 'language_code': language.short_name})
            rendition = template.render(context)
            outfile = os.path.join(TEMPLATES_STATICS_PATH, wizard.short_name + '_' + language.short_name + '.html')
            with open(outfile,'w') as o:
                o.write(rendition.encode('utf-8'))
            context = Context({'fields': [working_class.get_filtering_field()] + working_class.get_wizard_fields(), 'complete_fields': all_fields, 'language_code': language.short_name})
            rendition = template.render(context)
            outfile = os.path.join(TEMPLATES_STATICS_PATH, 'complete_' + wizard.short_name + '_' + language.short_name + '.html')
            with open(outfile,'w') as o:
                o.write(rendition.encode('utf-8'))
                
def generate_global_templates():
    languages = Attributes.objects.filter(active=True, type='available_language')
    template = loader.get_template('rendition/gso.html')
    for language in languages:
        entries = GlobalMenuEntries.objects.all().order_by('id')
        context = Context({'entries': entries, 'language_code': language.short_name})
        rendition = template.render(context)
        outfile = os.path.join(TEMPLATE_DIRS[0], 'gso_' + language.short_name + '.html')
        with open(outfile,'w') as o:
            o.write(rendition.encode('utf-8'))

def generate_attributes():
    all_types = Attributes.objects.all().order_by('type').distinct('type')
    languages = Attributes.objects.filter(active=True, type='available_language')
    for language in languages:
        context = Context({'selection': all_types, 'language_code': language.short_name})
        template = loader.get_template('rendition/attributes_list_option_renderer.html')
        rendition = template.render(context)
        outfile = os.path.join(TEMPLATES_STATICS_PATH, 'all_types_option_' + language.short_name + '.html')
        with open(outfile,'w') as o:
            o.write(rendition.encode('utf-8'))
        template = loader.get_template('rendition/attributes_list_select_renderer.html')
        rendition = template.render(context)
        outfile = os.path.join(TEMPLATES_STATICS_PATH, 'all_types_select_' + language.short_name + '.html')
        with open(outfile,'w') as o:
            o.write(rendition.encode('utf-8'))
        for a_type in all_types:
            all_elements = Attributes.objects.filter(type=a_type.type, active=True)
            context = Context({"selection": all_elements, 'language_code': language.short_name})
            template = loader.get_template('rendition/attributes_option_renderer.html')
            rendition = template.render(context)
            outfile = os.path.join(TEMPLATES_STATICS_PATH, a_type.type + '_' + language.short_name + '.html')
            with open(outfile,'w') as o:
                o.write(rendition.encode('utf-8'))
            template = loader.get_template('rendition/attributes_select_renderer.html')
            rendition = template.render(context)
            outfile = os.path.join(TEMPLATES_STATICS_PATH, a_type.type + '_select_' + language.short_name + '.html')
            with open(outfile,'w') as o:
                o.write(rendition.encode('utf-8'))
            template = loader.get_template('rendition/attributes_list_elements_renderer.html')
            rendition = template.render(context)
            outfile = os.path.join(TEMPLATES_STATICS_PATH, a_type.type + '_list_elements_' + language.short_name + '.html')
            with open(outfile,'w') as o:
                o.write(rendition.encode('utf-8'))

def populate_attributes_from_xlsx(model_name, xlsx_file):
    model = classes.my_class_import(model_name)
    workbook = load_workbook(xlsx_file)
    sheet = workbook.get_sheet_by_name(name=model.__name__)
    row_index = 1
    # Reading header
    header = []
    for column_index in range(1, sheet.get_highest_column() + 1):
        value = sheet.cell(row = row_index, column=column_index).value
        if value!=None:
            header.append(value if value!='' else header[-1])
        else:
            break
    LOGGER.info('Using header:' + str(header))
    row_index += 1
    while row_index<=sheet.get_highest_row():
        if model.objects.filter(identifier=sheet.cell(row = row_index, column=1).value).exists():
            instance = model.objects.get(identifier=sheet.cell(row = row_index, column=1).value)
        else:
            instance = model()
        for i in range(0,len(header)):
            value = sheet.cell(row = row_index, column=i+1).value
            setattr(instance, header[i], value)
        instance.save()
        row_index += 1
        
def populate_labels_from_xlsx(model_name, xlsx_file):
    model = classes.my_class_import(model_name)
    workbook = load_workbook(xlsx_file)
    sheet = workbook.get_sheet_by_name(name=model.__name__)
    row_index = 1
    # Reading header
    header = []
    for column_index in range(1, sheet.get_highest_column() + 1):
        value = sheet.cell(row = row_index, column=column_index).value
        if value!=None:
            header.append(value if value!='' else header[-1])
        else:
            break
    LOGGER.info('Using header:' + str(header))
    row_index += 1
    while row_index<=sheet.get_highest_row():
        if model.objects.filter(identifier=sheet.cell(row = row_index, column=1).value, language=sheet.cell(row = row_index, column=2)).exists():
            instance = model.objects.get(identifier=sheet.cell(row = row_index, column=1).value, language=sheet.cell(row = row_index, column=2))
        else:
            instance = model()
        for i in range(0,len(header)):
            value = sheet.cell(row = row_index, column=i+1).value
            setattr(instance, header[i], value)
        instance.save()
        row_index += 1

def populate_model_from_xlsx(model_name, xlsx_file):
    LOGGER.info("Loading data in " + model_name)
    model = classes.my_class_import(model_name)
    workbook = load_workbook(xlsx_file)
    sheet = workbook.get_sheet_by_name(name=model.__name__)
    row_index = 1
    # Reading header
    header = []
    for column_index in range(1, sheet.get_highest_column() + 1):
        value = sheet.cell(row = row_index, column=column_index).value
        if value!=None:
            header.append(value if value!='' else header[-1])
        else:
            break
    LOGGER.info('Using header:' + str(header))
    row_index += 1
    while row_index<=sheet.get_highest_row():
        instance = model()
        for i in range(0,len(header)):
            if sheet.cell(row = row_index, column=i+1).internal_value!=None:
                value = sheet.cell(row = row_index, column=i+1).value
                field_info = Attributes()
                field_info.short_name = header[i]
                field_info.name = header[i]
                instance.set_attribute('excel', field_info, value)
        instance.save()
        row_index += 1
        
class CoreModel(models.Model):

    def get_editable_fields(self):
        values = []
        for field in self.get_fields():
            if self._meta.get_field(field).get_internal_type()!='ManyToManyField':
                values.append(field)
        return values
        
    def get_associable_field(self):
        values = []
        for field in self.get_fields():
            if self._meta.get_field(field).get_internal_type()=='ManyToManyField':
                values.append(field)
        return values        

    @staticmethod
    def get_fields():
        return []
    
    def get_identifier(self):
        return 'name'
    
    def list_values(self):
        values = []
        for field in self.get_fields():
            LOGGER.debug(self.__class__.__name__ + ' * ' + field)
            if field in self._meta.get_all_field_names():
                if self._meta.get_field(field).get_internal_type()=='ManyToManyField' and getattr(self,field)!=None:
                    values.append(str([e.list_values() for e in list(getattr(self,field).all())]))
                elif self._meta.get_field(field).get_internal_type()=='ForeignKey' and getattr(self,field)!=None:
                    values.append(getattr(self,field).get_value())
                else:
                    values.append(str(getattr(self,field)))
            else:
                # Generic foreign key
                values.append(getattr(self,field).get_value())
        return values
    
    def get_value(self):
        if self.get_identifier()!=None:
            return getattr(self, self.get_identifier())
        else:
            return None
        
    def __unicode__(self):
        return unicode(self.get_value())
    
    def set_attribute(self, source, field_info, string_value):
        try:
            if string_value!='' and string_value!=None:
                if self._meta.get_field(field_info.short_name).get_internal_type()=='ManyToManyField':
                    if not self.many_fields.has_key(field_info.short_name):
                        self.many_fields[field_info.short_name] = []
                    foreign = self._meta.get_field(field_info.short_name).rel.to
                    foreign_element = foreign.retrieve_or_create(self, source, field_info.name, string_value)
                    if foreign_element!=None:                
                        self.many_fields[field_info.short_name].append(foreign_element)
                elif self._meta.get_field(field_info.short_name).get_internal_type()=='DateField' or self._meta.get_field(field_info.short_name).get_internal_type()=='DateTimeField':
                    try:
                        dt = datetime.datetime.strptime(string_value,'%m/%d/%Y' if source!='web' else '%y-%m-%d')
                        if self._meta.get_field(field_info.short_name).get_internal_type()=='DateField':
                            dt = datetime.date(dt.year, dt.month, dt.day)
                    except:
                        dt = string_value # This is not a String???
                    setattr(self, field_info.short_name, dt)
                elif self._meta.get_field(field_info.short_name).get_internal_type()=='ForeignKey':
                    linked_to = self._meta.get_field(field_info.short_name).rel.limit_choices_to
                    foreign = self._meta.get_field(field_info.short_name).rel.to
                    try:
                        filtering_by_id = dict(linked_to)
                        filtering_by_id['id'] = string_value
                        by_id = foreign.objects.filter(**filtering_by_id)
                    except ValueError:
                        by_id = None
                    try:
                        foreign._meta.get_field('name')
                        filtering_by_name = dict(linked_to)
                        filtering_by_name['name'] = string_value
                        by_name = foreign.objects.filter(**filtering_by_name)
                    except FieldDoesNotExist:
                        by_name = None
                    try:
                        foreign._meta.get_field('short_name')
                        filtering_by_short = dict(linked_to)
                        filtering_by_short['short_name'] = string_value
                        by_short = foreign.objects.filter(**filtering_by_short)
                    except FieldDoesNotExist:
                        by_short = None
                    if foreign==container.models.Attributes:
                        filtering_by_identifier = dict(linked_to)
                        filtering_by_identifier['identifier'] = string_value
                        by_identifier = foreign.objects.filter(**filtering_by_identifier)
                    else:
                        by_identifier = None
                    if by_id!=None and by_id.exists():
                        setattr(self, field_info.short_name, by_id[0])
                    elif by_name!=None and by_name.exists():
                        setattr(self, field_info.short_name, by_name[0])
                    elif by_short!=None and by_short.exists():
                        setattr(self, field_info.short_name, by_short[0])
                    elif by_identifier!=None and by_identifier.exists():
                        setattr(self, field_info.short_name, by_identifier[0])
                    else:
                        dict_entry = Dictionary.objects.filter(name=linked_to['type'], auto_create=True)
                        if dict_entry.exists():
                            LOGGER.info('Creating new attribute for ' + linked_to['type'] + ' with value ' + string_value)
                            new_attribute = Attributes()
                            new_attribute.active = True
                            new_attribute.identifier = dict_entry[0].identifier + str(string_value.upper()).replace(' ', '_')
                            new_attribute.name = string_value
                            new_attribute.short_name = string_value[0:32]
                            new_attribute.type = linked_to['type']
                            new_attribute.save()
                            setattr(self, field_info.short_name, new_attribute)
                        else:
                            LOGGER.warn('Cannot find foreign key instance on ' + str(self) + '.' + field_info.short_name + ' for value [' + string_value + '] and relation ' + str(linked_to))
                else:
                    if self._meta.get_field(field_info.short_name).get_internal_type()=='BooleanField':
                        string_value = string_value=='True'
                    setattr(self, field_info.short_name, string_value)
        except FieldDoesNotExist:
            #traceback.print_exc()
            LOGGER.error("Wrong field [" + field_info.short_name +"] for " + self.name)
    
    class Meta:
        ordering = ['id']

class FieldLabel(CoreModel):
    identifier = models.CharField(max_length=512)
    language = models.CharField(max_length=3)
    field_label = models.CharField(max_length=1024)
    
    @staticmethod
    def get_fields():
        return ['identifier','language','field_label']
    
    def get_identifier(self):
        return 'identifier'
    
class Attributes(CoreModel):
    identifier = models.CharField(max_length=128)
    name = models.CharField(max_length=128)
    short_name = models.CharField(max_length=32)
    type = models.CharField(max_length=64)
    active = models.BooleanField()
    
    @staticmethod
    def get_fields():
        return ['identifier','name','short_name','type','active']
    
    def get_short_json(self):
        return {'id': self.id, 'name': self.name, 'short_name': self.short_name, 'identifier': self.identifier }

    class Meta:
        ordering = ['name']
        

class MenuEntries(CoreModel):
    name = models.CharField(max_length=128)
    short_name = models.CharField(max_length=32)
    menu_target = models.ForeignKey(Attributes, limit_choices_to={'type':'container_menu_target'}, related_name='container_menu_target_rel', null=True)
    menu_type = models.CharField(max_length=128)
    container_type = models.ForeignKey(Attributes, limit_choices_to={'type':'container_type'}, related_name='container_type_menu_rel', null=True)
    language_identifier = models.CharField(max_length=64)
    data_target = models.CharField(max_length=128)
    action_type = models.CharField(max_length=128, null=True, blank=True)
    
    @staticmethod
    def get_fields():
        return ['menu_target','menu_type','container_type','language_identifier','data_target','action_type']

    class Meta:
        ordering = ['menu_target']
        
    
class GlobalMenuEntries(CoreModel):
    name = models.CharField(max_length=128)
    short_name = models.CharField(max_length=32)
    menu_target = models.CharField(max_length=128)
    menu_group_target = models.CharField(max_length=128, null=True, blank=True)
    menu_type = models.CharField(max_length=128)
    container_type = models.ForeignKey(Attributes, limit_choices_to={'type':'container_type'}, related_name='container_type_gbl_menu_rel', null=True)
    language_identifier = models.CharField(max_length=64)
    data_target = models.CharField(max_length=128)
    action_type = models.CharField(max_length=128, null=True, blank=True)
    administrator_only = models.BooleanField(default=False)

    @staticmethod
    def get_fields():
        return ['menu_target','menu_group_target','menu_type','container_type','language_identifier','data_target','action_type','administrator_only']

    class Meta:
        ordering = ['menu_target']
    
class Dictionary(CoreModel):
    identifier = models.CharField(max_length=128)
    name = models.CharField(max_length=128)
    auto_create = models.BooleanField()
    
    @staticmethod
    def get_fields():
        return ['identifier','name','auto_create']
    
    class Meta:
        ordering = ['name']
        
class Address(CoreModel):
    address_type = models.ForeignKey(Attributes, limit_choices_to={'type':'address_type'}, related_name='address_type_rel', null=True)
    line_1 = models.CharField(max_length=128, null=True)
    line_2 = models.CharField(max_length=128, null=True)
    zip_code = models.CharField(max_length=16, null=True)
    city = models.CharField(max_length=128, null=True)
    country = models.ForeignKey(Attributes, limit_choices_to={'type':'country_iso2'}, related_name='address_country_rel', null=True)

    def get_identifier(self):
        return 'id'

    @staticmethod
    def get_fields():
        return ['address_type','line_1','line_2','zip_code','city','country']
    
    @staticmethod
    def get_wizard_fields():
        return ['line_1','line_2','zip_code','city','country']
    
    @staticmethod
    def get_filtering_field():
        return "address_type"
    
    @staticmethod
    def get_displayed_fields(rendition_width):
        if rendition_width=='large':
            return ['address_type.identifier', 'line_1','line_2','zip_code','city', 'country.identifier']
        elif rendition_width=='small':
            return ['address_type.identifier', 'city', 'country.identifier']

    @staticmethod
    def retrieve_or_create(parent, source, key, value):
        if parent=='web':
            addresses = Address.objects.filter(address_type__identifier=value['address_type'], line_1=value['line_1'], line_2=value['line_2'], zip_code=value['zip_code'], city=value['city'], country__identifier=value['country'])
            if addresses.exists():
                new_address = addresses[0]
            else:
                new_address = Address()
                new_address.address_type = Attributes.objects.get(identifier=value['address_type'], active=True)
                new_address.line_1 = value['line_1']
                new_address.line_2 = value['line_2']
                new_address.zip_code = value['zip_code']
                new_address.city = value['city']
                new_address.country = Attributes.objects.get(identifier=value['country'], active=True)
                new_address.save()
            return new_address
        else:
            return None

class Email(CoreModel):
    address_type = models.ForeignKey(Attributes, limit_choices_to={'type':'email_type'}, related_name='email_type_rel', null=True)
    email_address = models.EmailField()

    def get_identifier(self):
        return 'id'

    @staticmethod
    def get_fields():
        return ['address_type','email_address']
    
    @staticmethod
    def get_wizard_fields():
        return ['email_address']
    
    @staticmethod
    def get_filtering_field():
        return "address_type"
    
    @staticmethod
    def get_displayed_fields(rendition_width):
        if rendition_width=='large':
            return ['address_type.identifier', 'email_address']
        elif rendition_width=='small':
            return ['address_type.identifier', 'email_address']
        
    @staticmethod
    def retrieve_or_create(parent, source, key, value):
        if parent=='web':
            emails = Email.objects.filter(address_type__identifier=value['address_type'], email_address=value['email_address'])
            if emails.exists():
                new_email = emails[0]
            else:
                new_email = Email()
                new_email.address_type = Attributes.objects.get(active=True, identifier=value['address_type'])
                new_email.email_address = value['email_address']
                new_email.save()
            return new_email
        else:
            return None
        
class Phone(CoreModel):
    line_type = models.ForeignKey(Attributes, limit_choices_to={'type':'phone_type'}, related_name='phone_type_rel', null=True)
    phone_number = models.TextField(max_length=32)

    @staticmethod
    def get_filtering_field():
        return "line_type"

    def get_identifier(self):
        return 'id'

    @staticmethod
    def get_fields():
        return ['line_type','phone_number']

    @staticmethod
    def get_wizard_fields():
        return ['phone_number']

    @staticmethod
    def get_displayed_fields(rendition_width):
        if rendition_width=='large':
            return ['line_type.identifier', 'phone_number']
        elif rendition_width=='small':
            return ['line_type.identifier', 'phone_number']
        
    @staticmethod
    def retrieve_or_create(parent, source, key, value):
        if parent=='web':
            emails = Email.objects.filter(address_type__identifier=value['line_type'], email_address=value['phone_number'])
            if emails.exists():
                new_phone = emails[0]
            else:
                new_phone = Phone()
                new_phone.line_type = Attributes.objects.get(active=True, identifier=value['line_type'])
                new_phone.phone_number = value['phone_number']
                new_phone.save()
            return new_phone
        else:
            return None
    
class Alias(CoreModel):
    alias_type = models.ForeignKey(Attributes, limit_choices_to={'type':'alias_type'}, related_name='alias_type_rel', null=True)
    alias_value = models.CharField(max_length=512)
    alias_additional = models.CharField(max_length=512)
    
    def get_identifier(self):
        return 'id'
    
    @staticmethod
    def get_fields():
        return ['alias_type','alias_value','alias_additional']

    @staticmethod
    def get_wizard_fields():
        return ['alias_value','alias_additional']

    @staticmethod
    def get_querying_fields():
        return ['alias_value','alias_additional']

    @staticmethod
    def get_filtering_field():
        return "alias_type"
    
    @staticmethod
    def get_displayed_fields(rendition_width):
        if rendition_width=='large':
            return ['alias_type.identifier','alias_value','alias_additional']
        elif rendition_width=='small':
            return ['alias_type.identifier','alias_value']
        
    @staticmethod
    def retrieve_or_create(parent, source, key, value):
        if isinstance(value, basestring):
            translation = Attributes.objects.filter(active=True, name=key, type=source.lower() + '_translation')
            if translation.exists():
                translation = translation[0].short_name
            else:
                translation = key
    
            alias_type = Attributes.objects.get(Q(active=True), Q(type='alias_type'), Q(identifier=translation) | Q(name=translation) | Q(short_name=translation))
            if not parent.aliases.filter(alias_type__id=alias_type.id).exists():
                new_alias = Alias()        
                new_alias.alias_type = alias_type
                new_alias.alias_value = value
                new_alias.alias_additional = ''
                new_alias.save()
                return new_alias
            else:
                return parent.aliases.get(alias_type__id=alias_type.id)
        else:
            if value['id']!=None and value['id']!='':
                alias = Alias.objects.get(id=value['id'])
            else:
                alias = Alias()
            for field in value.keys():
                if field not in ['id', 'many-to-many']:
                    if field!='alias_type':
                        setattr(alias, field, value[field])
                    else:
                        alias.alias_type = Attributes.objects.get(active=True, type='alias_type', identifier=value[field])
            alias.save()
            return alias

    class Meta:
        ordering = ['alias_value']

class Container(CoreModel):
    name = models.CharField(max_length=1024)
    short_name = models.CharField(max_length=512)
    type = models.ForeignKey(Attributes, limit_choices_to={'type':'container_type'}, related_name='container_type_rel', null=True)
    inception_date = models.DateField(null=True)
    closed_date = models.DateField(null=True)
    short_description = models.TextField(null=True, blank=True)
    status = models.ForeignKey(Attributes,limit_choices_to = {'type':'status'}, related_name='container_status_rel', null=True)
    
    many_fields = {}
    
    @staticmethod
    def get_wizard_fields():
        return ['name','short_name','inception_date','status']
    
    @classmethod
    def create(cls):
        entity = cls()
        entity.many_fields = {}
        return entity
    
    @staticmethod
    def get_fields():
        return ['name','short_name','type','inception_date','closed_date','status']
    
    @staticmethod
    def get_displayed_fields(rendition_width):
        if rendition_width=='large':
            return ['name', 'short_name', 'type','status']
        elif rendition_width=='small':
            return ['name', 'type']

    def finalize(self):
        active = Attributes.objects.get(active=True, type='status', identifier='STATUS_ACTIVE')
        self.status = active
        self.save()
        # TODO Loop on many to many only
        for field_name in self._meta.get_all_field_names():
            try:
                if self._meta.get_field(field_name).get_internal_type()=='ManyToManyField':
                    if self.many_fields.has_key(field_name):
                        values = list(self.many_fields[field_name])
                        setattr(self, field_name, values)
            except FieldDoesNotExist:
                None
        self.save()
        
    @staticmethod
    def get_querying_fields():
        return ['name', 'short_name', 'short_description']
        
    @staticmethod
    def get_filtering_field():
        return "type"
    
    def clean(self):
        self.many_fields = {}
        for field_name in self._meta.get_all_field_names():
            try:
                if self._meta.get_field(field_name).get_internal_type()=='ForeignKey':
                    setattr(self, field_name, None)
                elif self._meta.get_field(field_name).get_internal_type()=='ManyToManyField':
                    setattr(self, field_name, [])
            except FieldDoesNotExist:
                None

    def clone(self, _class):
        _clone = _class()
        [setattr(_clone, fld.name, getattr(self, fld.name)) for fld in self._meta.fields if fld.name != self._meta.pk and fld.name.find('_ptr')==-1]
        _clone.id = None
        _clone.pk = None
        return _clone

    class Meta:
        ordering = ['name']
        
class ContainerDocument(Container):
    containers = models.ManyToManyField("Container", related_name='container_document_rel')
    last_uploader = models.ForeignKey(User, related_name='document_uploader_rel')
        
class Universe(Container):
    public = models.BooleanField()
    members = models.ManyToManyField("Container", related_name='universe_inventory_rel')
    owner = models.ForeignKey('ThirdPartyContainer', related_name='universe_owner_rel')
    description = models.TextField(null=True, blank=True)
    
    
class ThirdPartyContainer(Container):
    addresses = models.ManyToManyField(Address)
    emails = models.ManyToManyField(Email)
    phones = models.ManyToManyField(Phone)

    @staticmethod
    def get_fields():
        return Container.get_fields() + ['addresses','emails','phones']

class PersonContainer(ThirdPartyContainer):
    first_name = models.CharField(max_length=128)
    last_name = models.CharField(max_length=128)
    birth_date = models.DateField(null=True)
    
    @staticmethod
    def get_fields():
        return ThirdPartyContainer.get_fields() + ['first_name','last_name','birth_date']
    
    def get_short_json(self):
        return {'id': self.id, 'first_name': self.first_name, 'last_name': self.last_name}
    
    @staticmethod
    def get_querying_fields():
        return ['name', 'short_name', 'short_description', 'first_name', 'last_name']

class CompanyMember(CoreModel):
    person = models.ForeignKey(PersonContainer, null=True)
    role = models.ForeignKey(Attributes, limit_choices_to={'type':'company_member_role'}, related_name='company_member_role_rel', null=True)

    def get_identifier(self):
        return 'id'

    @staticmethod
    def get_fields():
        return ['person','role']
    
    @staticmethod
    def get_filtering_field():
        return "role"
    
    @staticmethod
    def get_querying_class():
        return PersonContainer
    
    @staticmethod
    def get_displayed_fields(rendition_width):
        if rendition_width=='large':
            return ['role.name', 'person.last_name', 'person.first_name', 'person.birth_date']
        elif rendition_width=='small':
            return ['role.name', 'person.name']
        
    @staticmethod
    def get_wizard_fields():
        return ['person']
    
    @staticmethod
    def retrieve_or_create(parent, source, key, value):
        if parent=='web' and source==None and key==None and isinstance(value, dict):
            print value
            if value['id']!=None and value['id']!='':
                member = CompanyMember.objects.get(id=value['id'])
            else:
                member = CompanyMember()
            member.role = Attributes.objects.get(active=True, type='company_member_role', identifier=value['role'])
            member.person = PersonContainer.objects.get(name=value['person'])
            member.save()
            return member
            

class CompanyContainer(ThirdPartyContainer):
    members = models.ManyToManyField(CompanyMember)
    subsidiary = models.ManyToManyField('CompanySubsidiary')
    
    @staticmethod
    def get_fields():
        return ThirdPartyContainer.get_fields() + ['members','subsidiary']

    def get_short_json(self):
        data_provider = Attributes.objects.get(identifier='SCR_DP', active=True)
        is_provider = RelatedCompany.objects.filter(company=self, role=data_provider).exists()

        return {'id': self.id, 'name': self.name, 'short_name': self.short_name, 'provider': is_provider}
    
    @staticmethod
    def get_querying_fields():
        return ['name', 'short_name']

class CompanySubsidiary(CoreModel):
    company = models.ForeignKey('CompanyContainer', null=True)
    role = models.ForeignKey(Attributes, limit_choices_to={'type':'company_subsidiary_role'}, related_name='company_subsidiary_role_rel', null=True)

    def get_identifier(self):
        return 'id'
    
    @staticmethod
    def get_fields():
        return ['company','role']
    
    @staticmethod
    def get_filtering_field():
        return "role"
    
    @staticmethod
    def get_querying_class():
        return CompanyContainer
    
    @staticmethod
    def get_displayed_fields(rendition_width):
        if rendition_width=='large':
            return ['company.name', 'company.inception_date', 'company.status.name']
        elif rendition_width=='small':
            return ['company.name']
        
    @staticmethod
    def get_wizard_fields():
        return ['company']
    
    @staticmethod
    def retrieve_or_create(parent, source, key, value):
        if parent=='web' and source==None and key==None and isinstance(value, dict):
            print value
            if value['id']!=None and value['id']!='':
                subsidiary = CompanySubsidiary.objects.get(id=value['id'])
            else:
                subsidiary = CompanySubsidiary()
            subsidiary.role = Attributes.objects.get(active=True, type='company_subsidiary_role', identifier=value['role'])
            subsidiary.company = CompanyContainer.objects.get(name=value['company'])
            subsidiary.save()
            return subsidiary
    
class RelatedCompany(CoreModel):
    company = models.ForeignKey(CompanyContainer, null=True)
    role = models.ForeignKey(Attributes, limit_choices_to={'type':'security_company_role'}, related_name='security_company_role_rel', null=True)

    def get_identifier(self):
        return 'id'

    @staticmethod
    def get_fields():
        return ['company','role']

    @staticmethod
    def get_displayed_fields(rendition_width):
        if rendition_width=='large':
            return ['company.name','role.name','company.status.name']
        elif rendition_width=='small':
            return ['company.name','role.name']
    
    @staticmethod
    def get_filtering_field():
        return "role"
    
    @staticmethod
    def get_querying_class():
        return CompanyContainer
    
    @staticmethod
    def retrieve_or_create(parent, source, key, value):
        if parent!='web':
            translation = Attributes.objects.filter(active=True, name=key, type=source.lower() + '_translation')
            if translation.exists():
                translation = translation[0].short_name
            else:
                translation = key
            working_value = value
        else:
            translation = value['role']
            working_value = value['company']
            
        company = CompanyContainer.objects.filter(name=working_value)
        if company.exists():
            company = company[0]
        else:
            company = CompanyContainer()
            company.name = working_value
            company.short_name = 'Company...'
            tbv = Attributes.objects.get(active=True, type='status', identifier='STATUS_TO_BE_VALIDATED')
            company.status = tbv
            company.type = Attributes.objects.get(active=True, type='container_type', identifier='CONT_COMPANY')
            company.save()
        role = Attributes.objects.get(Q(active=True), Q(type='security_company_role'), Q(identifier=translation) | Q(name=translation) | Q(short_name=translation))
        relation = RelatedCompany.objects.filter(company__id=company.id, role__id=role.id)
        if relation.exists():
            return relation[0]
        else:
            new_relation = RelatedCompany()        
            new_relation.role = role
            new_relation.company = company
            new_relation.save()
            return new_relation


class RelatedThird(CoreModel):
    third = models.ForeignKey(PersonContainer)
    role = models.ForeignKey(Attributes, limit_choices_to={'type':'security_third_role'}, related_name='security_third_role_rel', null=True)

    def get_identifier(self):
        return 'id'

    @staticmethod
    def get_fields():
        return ['third','role']
    
    @staticmethod
    def get_filtering_field():
        return "role"
    
    @staticmethod
    def get_querying_class():
        return PersonContainer
    
    @staticmethod
    def get_displayed_fields(rendition_width):
        if rendition_width=='large':
            return ['third.name','role.name','third.status.name']
        elif rendition_width=='small':
            return ['third.name','role.name']
    
    @staticmethod
    def retrieve_or_create(parent, source, key, value):
        translation = Attributes.objects.filter(active=True, name=key, type=source.lower() + '_translation')
        if translation.exists():
            translation = translation[0].short_name
        else:
            translation = key
        third = PersonContainer.objects.filter(name=value)
        if third.exists():
            third = third[0]
        else:
            third = PersonContainer()
            third.name = value
            third.short_name = 'Company...'
            tbv = Attributes.objects.get(active=True, type='status', identifier='STATUS_TO_BE_VALIDATED')
            third.status = tbv
            third.type = Attributes.objects.get(active=True, type='container_type', identifier='CONT_PERSON')
            third.save()
            
        new_relation = RelatedThird()        
        new_relation.role = Attributes.objects.get(Q(active=True), Q(type='security_third_role'), Q(identifier=translation) | Q(name=translation) | Q(short_name=translation))
        new_relation.third = third
        new_relation.save()
        return new_relation
    
class UserMapping(models.Model):
    third_container = models.ForeignKey(ThirdPartyContainer, related_name='user_related_third', null=False)
    related_user = models.ForeignKey(User, related_name='related_third_user', null=False)